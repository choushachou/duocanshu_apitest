#coding:utf - 8
import os
path = os.path.dirname(os.path.abspath(__file__))#获得上一级路径
from openpyxl import *
import requests
from api_tests.util.json_util import OperationJson
import json

class OperationExcel:
    def __init__(self,file_name=None,sheet_id=None):
        self.file_name=path+'/ceshi.xlsx'
        

    def get_data(self): #传入参数，打开表格，获取表格内容
        data = load_workbook(self.file_name)
        tables = data['设备-245']
        return tables

    def get_rowss(self):
        #获取表格最大行数
        tables=self.get_data()
        return tables.max_row

    #获取表格最大列数
    def get_liness(self):
        tables = self.get_data()
        return tables.max_column

    def get_value(self,row,col):
        #传入行和列，获取某个单元格的内容
        return self.get_data().cell(row,col).value

    # 获取某列的内容
    def getColValues(self, col):
        rows = self.get_rowss()
        values = []

        for i in range(2, rows + 1):
            col_value = self.get_value(i, col)
            values.append(col_value)
        return values

    # 获取某行的内容
    def getRowValues(self, row):
        cols = self.get_liness()
        values = []
        for i in range(1, cols + 1):
            row_value = self.get_value(row, i)
            values.append(row_value)
        return values

    def write_value(self,row,col,value):
        #在表格中输入内容
        r_data=load_workbook(filename=self.file_name)
        sheet_data=r_data['设备-245']

        sheet_data.cell(row,col,value)  #写内容
        r_data.save(self.file_name)  #保存

#下面为依赖的内容

    def get_onecol_data(self,col_id=None):
        #huoqu第一列的内容
        if col_id != None:
            cols = self.getColValues(col_id)
        else:
            cols = self.getColValues(1)
        return cols
    def get_row_num(self,case_id):
        #gen据对应的caseid找到对应的行号
        num = 2
        cols_data = self.get_onecol_data()
        for col_data in cols_data:
            if str(case_id) in str(col_data):
                return num
            num = num +1


    def get_rows_data(self,case_id):
        #根据对应的caseID找到对应的内容
        row_num=self.get_row_num(case_id)
        tables = self.get_data()
        row_data = self.getRowValues(row_num)
        return row_data


    #请求头
    def h_token(self,row,col):
        url = str(self.get_data().cell(row,col).value)
        data1 = json.loads(self.get_data().cell(row,col+1).value)
        
        # method = self.get_data().cell(row,col-1).value
        # #res_method = RunMethod.run_main()
        # res_a = self.res_method(method=method,url=url,data=data).json()
        res_a = requests.post(url=url,data=data1)
        code= res_a.status_code
        res_a = res_a.json()
        res_b = res_a["data"]#获得token值
        header = res_b["token"]
        headers = {"token": header}
        op_json= OperationJson()
        op_json.write_data(headers)
        return res_a,code
    # def write_cookie(self,row,col):
    #     #把token写道json文件中
    #     cookie = self.h_token(row=row,col=col)

if __name__ == '__main__':
    s=OperationExcel().get_rowss()
    print(s)